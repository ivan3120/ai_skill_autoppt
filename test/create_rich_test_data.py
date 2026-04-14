#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""创建丰富的测试数据"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import random
from datetime import datetime, timedelta

wb = openpyxl.Workbook()

# 删除默认sheet
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# ==================== 整体评估概览 ====================
ws_overview = wb.create_sheet('整体评估概览', 0)
ws_overview.merge_cells('A1:D1')
ws_overview['A1'] = '云核心网网络架构评估报告'
ws_overview['A1'].font = Font(size=16, bold=True)
ws_overview['A1'].alignment = Alignment(horizontal='center')

# 产品域数据
ws_overview['A3'] = '产品域'
ws_overview['B3'] = '网元数量'
ws_overview['C3'] = '容灾组数'
ws_overview['D3'] = '区域数'

domains_data = [
    ('5G核心网', 85, 12, 8),
    ('��联网专用', 45, 6, 4),
    ('VoNR核心网', 32, 4, 3),
    ('MEC平台', 18, 3, 2),
    ('短信中心', 15, 2, 2),
    ('VoLTE', 28, 4, 3),
    ('计费中心', 12, 2, 1),
]

for col in ['A3', 'B3', 'C3', 'D3']:
    ws_overview[col].font = Font(bold=True, color='FFFFFF')
    ws_overview[col].fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')

for idx, (domain, ne_count, dr_count, area_count) in enumerate(domains_data, start=4):
    ws_overview.cell(row=idx, column=1, value=domain)
    ws_overview.cell(row=idx, column=2, value=ne_count)
    ws_overview.cell(row=idx, column=3, value=dr_count)
    ws_overview.cell(row=idx, column=4, value=area_count)

# 评估维度汇总
ws_overview['A12'] = '评估维度'
ws_overview['B12'] = '通过率'
ws_overview['C12'] = '风险等级'
ws_overview['D12'] = '主要问题'

dimensions_data = [
    ('MT0-网元高稳', '84%', '低', '无'),
    ('MT1-部署离散度', '73%', '中', '部分网元未实现同城双活'),
    ('MT2-组网架构高可用', '88%', '低', '无'),
    ('MT3-业务路由高可用', '79%', '中', '部分业务路由单一'),
    ('MT4-网络容量均衡度', '60%', '高', '负载不均衡,部分网元负载超过80%'),
    ('MT5-网元版本EOS', '92%', '低', '无'),
    ('MT6-云平台版本EOS', '85%', '低', '无'),
]

for col in ['A12', 'B12', 'C12', 'D12']:
    ws_overview[col].font = Font(bold=True)

for idx, row in enumerate(dimensions_data, start=13):
    for col_idx, value in enumerate(row, start=1):
        ws_overview.cell(row=idx, column=col_idx, value=value)

# ==================== MT0-网元高稳评估 ====================
ws_mt0 = wb.create_sheet('MT0-网元高稳评估')
ws_mt0['A1'] = '网元名称'
ws_mt0['B1'] = '高稳类型'
ws_mt0['C1'] = '状态'
ws_mt0['D1'] = '备注'

ne_types = ['N+1备份', '同城双活', '异地容灾', '集群部署']
nes = [
    ('AMF01', 'N+1备份', '正常', ''),
    ('AMF02', 'N+1备份', '正常', ''),
    ('SMF01', '同城双活', '正常', ''),
    ('SMF02', '同城双活', '正常', ''),
    ('UPF01', '集群部署', '正常', ''),
    ('UPF02', '集群部署', '正常', ''),
    ('UPF03', 'N+1备份', '正常', ''),
    ('AUSF01', 'N+1备份', '正常', ''),
    ('AUSF02', 'N+1备份', '正常', ''),
    ('NRF01', '同城双活', '正常', ''),
    ('NRF02', '同城双活', '正常', ''),
    ('NSSF01', 'N+1备份', '正常', ''),
    ('PCF01', '同城双活', '正常', ''),
    ('PCF02', 'N+1备份', '正常', ''),
    ('UDM01', '同城双活', '正常', ''),
    ('UDR01', 'N+1备份', '正常', ''),
    ('CHF01', 'N+1备份', '正常', ''),
    ('CHF02', 'N+1备份', '正常', ''),
    ('SMSF01', 'N+1备份', '正常', ''),
    ('MME01', 'N+1备份', '正常', '4G网元'),
    ('SGW01', '集群部署', '正常', ''),
    ('PGW01', '集群部署', '正常', ''),
    ('PCRF01', 'N+1备份', '正常', ''),
    ('BOSS01', '同城双活', '正常', '计费系统'),
]

for col in ['A1', 'B1', 'C1', 'D1']:
    ws_mt0[col].font = Font(bold=True)

for idx, (ne, ht, st, note) in enumerate(nes, start=2):
    ws_mt0.cell(row=idx, column=1, value=ne)
    ws_mt0.cell(row=idx, column=2, value=ht)
    ws_mt0.cell(row=idx, column=3, value=st)
    ws_mt0.cell(row=idx, column=4, value=note)

# ==================== MT1-部署离散度评估 ====================
ws_mt1 = wb.create_sheet('MT1-部署离散度评估')
ws_mt1['A1'] = '网元名称'
ws_mt1['B1'] = 'AZ1'
ws_mt1['C1'] = 'AZ2'
ws_mt1['D1'] = '离散度得分'
ws_mt1['E1'] = '评估结果'

for col in ['A1', 'B1', 'C1', 'D1', 'E1']:
    ws_mt1[col].font = Font(bold=True)

mt1_data = [
    ('AMF01', 'AZ1', 'AZ2', 85, '通过'),
    ('AMF02', 'AZ1', 'AZ2', 90, '通过'),
    ('SMF01', 'AZ1', 'AZ2', 88, '通过'),
    ('SMF02', 'AZ1', 'AZ2', 82, '通过'),
    ('UPF01', 'AZ1', 'AZ2', 75, '通过'),
    ('UPF02', 'AZ1', 'AZ3', 65, '待优化'),
    ('UPF03', 'AZ1', '', 40, '不通过'),
    ('AUSF01', 'AZ1', 'AZ2', 78, '通过'),
    ('NRF01', 'AZ1', 'AZ2', 92, '通过'),
    ('NSSF01', 'AZ1', 'AZ2', 80, '通过'),
    ('PCF01', 'AZ1', 'AZ2', 85, '通过'),
    ('UDM01', 'AZ1', 'AZ2', 88, '通过'),
    ('CHF01', 'AZ1', 'AZ2', 72, '待优化'),
    ('MME01', 'AZ1', 'AZ2', 70, '待优化'),
    ('SGW01', 'AZ1', 'AZ2', 68, '待优化'),
]

for idx, row in enumerate(mt1_data, start=2):
    for col_idx, value in enumerate(row, start=1):
        ws_mt1.cell(row=idx, column=col_idx, value=value)

# ==================== MT2-组网架构高可用 ====================
ws_mt2 = wb.create_sheet('MT2-组网架构高可用')
ws_mt2['A1'] = '网元对'
ws_mt2['B1'] = '组网方式'
ws_mt2['C1'] = '高可用状态'
ws_mt2['D1'] = '评估结果'

for col in ['A1', 'B1', 'C1', 'D1']:
    ws_mt2[col].font = Font(bold=True)

mt2_data = [
    ('AMF01/AMF02', '主备', '正常', '通过'),
    ('SMF01/SMF02', '主备', '正常', '通过'),
    ('UPF01/UPF02', '集群', '正常', '通过'),
    ('NRF01/NRF02', '主备', '正常', '通过'),
    ('PCF01/PCF02', '主备', '正常', '通过'),
    ('UDM01/UDM02', '主备', '正常', '通过'),
    ('CHF01/CHF02', '主备', '正常', '通过'),
    ('SGW01/SGW02', '集群', '正常', '通过'),
]

for idx, row in enumerate(mt2_data, start=2):
    for col_idx, value in enumerate(row, start=1):
        ws_mt2.cell(row=idx, column=col_idx, value=value)

# ==================== MT3-业务路由高可用 ====================
ws_mt3 = wb.create_sheet('MT3-业务路由高可用')
ws_mt3['A1'] = '业务名称'
ws_mt3['B1'] = '主用路由'
ws_mt3['C1'] = '备用路由'
ws_mt3['D1'] = '评估结果'

for col in ['A1', 'B1', 'C1', 'D1']:
    ws_mt3[col].font = Font(bold=True)

mt3_data = [
    ('5G注册', 'AMF01->AUSF01', 'AMF02->AUSF02', '通过'),
    ('PDU会话', 'SMF01->UPF01', 'SMF02->UPF02', '通过'),
    ('VoNR呼叫', 'AMF01->SMF01', 'AMF02->SMF02', '通过'),
    ('短信业务', 'CHF01->SMSF01', '', '待优化'),
    ('物联网', 'SMF03->UPF03', '', '待优化'),
    ('漫游业务', 'SMF04->UPF04', 'SMF05->UPF05', '通过'),
    ('计费话单', 'CHF01->BOSS01', 'CHF02->BOSS02', '通过'),
]

for idx, row in enumerate(mt3_data, start=2):
    for col_idx, value in enumerate(row, start=1):
        ws_mt3.cell(row=idx, column=col_idx, value=value)

# ==================== MT4-网络容量均衡度 ====================
ws_mt4 = wb.create_sheet('MT4-网络容量-网元间均衡度')
ws_mt4['A1'] = '网元名称'
ws_mt4['B1'] = '当前负载'
ws_mt4['C1'] = '容量上限'
ws_mt4['D1'] = '负载率'
ws_mt4['E1'] = '评估结果'

for col in ['A1', 'B1', 'C1', 'D1', 'E1']:
    ws_mt4[col].font = Font(bold=True)

mt4_data = [
    ('AMF01', 120000, 150000, '80%', '高'),
    ('AMF02', 95000, 150000, '63%', '正常'),
    ('SMF01', 85000, 100000, '85%', '高'),
    ('SMF02', 72000, 100000, '72%', '中'),
    ('UPF01', 450000, 500000, '90%', '高'),
    ('UPF02', 380000, 500000, '76%', '中'),
    ('UPF03', 520000, 500000, '104%', '过载'),
    ('NRF01', 45000, 100000, '45%', '正常'),
    ('PCF01', 28000, 50000, '56%', '正常'),
    ('UDM01', 52000, 80000, '65%', '正常'),
    ('CHF01', 85000, 100000, '85%', '高'),
    ('CHF02', 68000, 100000, '68%', '正常'),
    ('SGW01', 320000, 400000, '80%', '高'),
    ('PGW01', 280000, 400000, '70%', '正常'),
]

for idx, row in enumerate(mt4_data, start=2):
    for col_idx, value in enumerate(row, start=1):
        ws_mt4.cell(row=idx, column=col_idx, value=value)

# ==================== MT5-网元版本EOS ====================
ws_mt5 = wb.create_sheet('MT5-网元版本EOS')
ws_mt5['A1'] = '网元名称'
ws_mt5['B1'] = '当前版本'
ws_mt5['C1'] = 'EOS日期'
ws_mt5['D1'] = '剩余天数'
ws_mt5['E1'] = '评估结果'

for col in ['A1', 'B1', 'C1', 'D1', 'E1']:
    ws_mt5[col].font = Font(bold=True)

mt5_data = [
    ('AMF01', 'V3.5.0', '2026-12-31', 630, '正常'),
    ('SMF01', 'V3.2.0', '2027-03-31', 720, '正常'),
    ('UPF01', 'V4.0.0', '2027-06-30', 810, '正常'),
    ('NRF01', 'V2.8.0', '2025-06-30', 30, '预警'),
    ('PCF01', 'V3.0.0', '2025-12-31', 260, '正常'),
    ('UDM01', 'V3.1.0', '2026-03-31', 350, '正常'),
    ('CHF01', 'V2.9.0', '2024-12-31', -100, '过期'),
    ('MME01', 'V1.5.0', '2024-06-30', -280, '过期'),
]

for idx, row in enumerate(mt5_data, start=2):
    for col_idx, value in enumerate(row, start=1):
        ws_mt5.cell(row=idx, column=col_idx, value=value)

# ==================== MT6-云平台版本EOS ====================
ws_mt6 = wb.create_sheet('MT6-云平台版本EOS')
ws_mt6['A1'] = '平台名称'
ws_mt6['B1'] = '当前版本'
ws_mt6['C1'] = 'EOS日期'
ws_mt6['D1'] = '剩余天数'
ws_mt6['E1'] = '评估结果'

for col in ['A1', 'B1', 'C1', 'D1', 'E1']:
    ws_mt6[col].font = Font(bold=True)

mt6_data = [
    ('OpenStack-Victoria', 'V20.12', '2025-06-30', 30, '预警'),
    ('OpenStack-Wallaby', 'V21.12', '2026-06-30', 400, '正常'),
    ('Kubernetes-V1.21', 'V1.21.0', '2024-06-30', -280, '过期'),
    ('Kubernetes-V1.25', 'V1.25.0', '2025-12-31', 260, '正常'),
    ('Kubernetes-V1.28', 'V1.28.0', '2026-12-31', 630, '正常'),
    ('VMware-V7.0', 'V7.0.0', '2027-03-31', 720, '正常'),
]

for idx, row in enumerate(mt6_data, start=2):
    for col_idx, value in enumerate(row, start=1):
        ws_mt6.cell(row=idx, column=col_idx, value=value)

# ==================== 评估维度说明 ====================
ws_desc = wb.create_sheet('评估维度说明')
ws_desc['A1'] = '维度'
ws_desc['B1'] = '说明'
ws_desc['C1'] = '权重'

dimensions_desc = [
    ('MT0', '网元高稳评估', '20%'),
    ('MT1', '部署离散度评估', '15%'),
    ('MT2', '组网架构高可用评估', '15%'),
    ('MT3', '业务路由高可用评估', '15%'),
    ('MT4', '网络容量均衡度评估', '20%'),
    ('MT5', '网元版本EOS评估', '10%'),
    ('MT6', '云平台版本EOS评估', '5%'),
]

for col in ['A1', 'B1', 'C1']:
    ws_desc[col].font = Font(bold=True)

for idx, row in enumerate(dimensions_desc, start=2):
    for col_idx, value in enumerate(row, start=1):
        ws_desc.cell(row=idx, column=col_idx, value=value)

# 保存文件
wb.save('assets/test_operator_A_5G_rich.xlsx')
print('Created: assets/test_operator_A_5G_rich.xlsx')
print(f'Sheets: {wb.sheetnames}')